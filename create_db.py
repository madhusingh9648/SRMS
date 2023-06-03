import sqlite3
from openpyxl import Workbook

def create_db():
    con = sqlite3.connect(database="rms.db")
    cur = con.cursor()
    cur.execute("CREATE TABLE IF NOT EXISTS course(cid INTEGER PRIMARY KEY AUTOINCREMENT,name text,faculty text,year text,description text)")
    con.commit()

    cur.execute("CREATE TABLE IF NOT EXISTS student(roll INTEGER PRIMARY KEY AUTOINCREMENT,name text,email text,gender text,dob text,contact text,admission text,course text,state text,city text,pin text,address text)")
    con.commit()

    cur.execute("CREATE TABLE IF NOT EXISTS result(rid INTEGER PRIMARY KEY AUTOINCREMENT,roll text,name text,course text,marks_ob text,full_marks text,per text)")
    con.commit()

    cur.execute("CREATE TABLE IF NOT EXISTS employee(eid INTEGER PRIMARY KEY AUTOINCREMENT,f_name text,l_name text,contact text,email text,question text,answer text,password text)")
    con.commit()

    # Retrieve data from the tables
    cur.execute("SELECT * FROM course")
    course_data = cur.fetchall()

    cur.execute("SELECT * FROM student")
    student_data = cur.fetchall()

    cur.execute("SELECT * FROM result")
    result_data = cur.fetchall()

    cur.execute("SELECT * FROM employee")
    employee_data = cur.fetchall()

    con.close()

    # Write data to Excel file
    wb = Workbook()
    course_sheet = wb.active
    course_sheet.title = "Course Data"
    course_sheet.append(("Course ID", "Name", "Faculty", "Year", "Description"))
    for row in course_data:
        course_sheet.append(row)

    student_sheet = wb.create_sheet(title="Student Data")
    student_sheet.append(("Roll", "Name", "Email", "Gender", "DOB", "Contact", "Admission Date", "Course", "State", "City", "PIN", "Address"))
    for row in student_data:
        student_sheet.append(row)

    result_sheet = wb.create_sheet(title="Result Data")
    result_sheet.append(("Result ID", "Roll", "Name", "Course", "Marks Obtained", "Full Marks", "Percentage"))
    for row in result_data:
        result_sheet.append(row)

    employee_sheet = wb.create_sheet(title="Employee Data")
    employee_sheet.append(("Employee ID", "First Name", "Last Name", "Contact", "Email", "Question", "Answer", "Password"))
    for row in employee_data:
        employee_sheet.append(row)

    wb.save("rms_data.xlsx")

create_db()
